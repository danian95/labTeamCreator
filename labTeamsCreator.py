import os
import random
import easygui
import openpyxl
from copy import copy
import pickle
import sys

def read_excel(path) -> [str]:
	workbook = openpyxl.load_workbook(path, data_only = True)
	worksheet = workbook.worksheets[0]
	students = []
	for i in range(2,102):
		lastName = str(worksheet.cell(row = i, column = 1).value)
		firstName = str(worksheet.cell(row = i, column = 2).value)
		if lastName != '' and lastName != 'None':
			students.append(lastName + ' ' + firstName)
		else:
			break
	return students

def read_teams(path,teamNb) -> ([[str]],[str]):
	if not os.path.exists(path):
		return [],[]
	workbook = openpyxl.load_workbook(path)
	if len(workbook.worksheets) < teamNb:
		return [],[]
	worksheet = workbook.worksheets[teamNb - 1]
	teams = []
	allStudents = []
	for i in range(2,101):
		firstStudent = str(worksheet.cell(row = i, column = 2).value)
		if firstStudent == '' or firstStudent == 'None':
			break
		secondStudent = str(worksheet.cell(row=i, column=3).value)
		if secondStudent == '' or secondStudent == 'None':
			raise ValueError('Invalid team in retrieved lab ' + os.path.basename(path) + ': only found 1 student in the team.')
		thirdStudent = str(worksheet.cell(row=i, column=4).value)
		if thirdStudent == '' or thirdStudent == 'None':
			teams.append([firstStudent,secondStudent])
			allStudents.append(firstStudent)
			allStudents.append(secondStudent)
		else:
			teams.append([firstStudent,secondStudent,thirdStudent])
			allStudents.append(firstStudent)
			allStudents.append(secondStudent)
			allStudents.append(thirdStudent)
	return teams,allStudents

def copyCell(destCell,orgCell,value):
	destCell.font = copy(orgCell.font)
	destCell.border = copy(orgCell.border)
	destCell.fill = copy(orgCell.fill)
	destCell.number_format = copy(orgCell.number_format)
	destCell.protection = copy(orgCell.protection)
	destCell.alignment = copy(orgCell.alignment)
	destCell.value = value
def write_template(teams,sheet,labNb):
	try:
		templatePath = os.path.join(sys._MEIPASS,'template.xlsx')
	except Exception:
		templatePath = 'template.xlsx'
	template = openpyxl.load_workbook(templatePath).worksheets[0]
	rowNb = len(teams)
	colNb = 2
	for team in teams:
		if len(team) == 3:
			colNb = 3

	sheet.merge_cells(start_row=6, start_column=8, end_row=6, end_column=8 + colNb)
	sheet.merge_cells(start_row=7, start_column=8, end_row=7, end_column=8 + colNb)
	sheet.cell(6,8).value = 'Laboratoire ' + str(labNb)
	sheet.cell(6,8).alignment = copy(template.cell(1,1).alignment)
	copyCell(sheet.cell(8,8),template.cell(1,1),'Équipe')
	copyCell(sheet.cell(9, 8), template.cell(2, 1), 1)
	copyCell(sheet.cell(8,9), template.cell(1, 2), 'Nom 1')
	copyCell(sheet.cell(9, 9), template.cell(2, 2), teams[0][0])
	sheet.column_dimensions[sheet.cell(1, 9).column_letter].width = 25
	if colNb == 3:
		thirdStudent =  teams[0][2] if len(teams[0]) == 3 else ''
		copyCell(sheet.cell(8, 10), template.cell(1, 2), 'Nom 2')
		copyCell(sheet.cell(9, 10), template.cell(2, 2), teams[0][1])
		sheet.column_dimensions[sheet.cell(1, 10).column_letter].width = 25
		copyCell(sheet.cell(8, 11), template.cell(1, 3), 'Nom 3')
		copyCell(sheet.cell(9, 11), template.cell(2, 3), thirdStudent)
		sheet.column_dimensions[sheet.cell(1, 11).column_letter].width = 25
	else:
		copyCell(sheet.cell(8, 10), template.cell(1, 3), 'Nom 2')
		copyCell(sheet.cell(9, 10), template.cell(2, 3), teams[0][1])
		sheet.column_dimensions[sheet.cell(1, 10).column_letter].width = 25
	for r in range(2,rowNb):
		if r % 2 == 0:
			tr = 3
		else:
			tr = 4
		team = teams[r-1]
		copyCell(sheet.cell(8+r, 8), template.cell(tr, 1), r)
		copyCell(sheet.cell(8+r, 9), template.cell(tr, 2), team[0])
		if colNb == 3:
			thirdStudent = team[2] if len(team) == 3 else ''
			copyCell(sheet.cell(8+r, 10), template.cell(tr, 2), team[1])
			copyCell(sheet.cell(8+r, 11), template.cell(tr, 3), thirdStudent)
		else:
			copyCell(sheet.cell(8+r, 10), template.cell(tr, 3), team[1])

	if rowNb % 2 == 0:
		tr = 5
	else:
		tr = 6
	lastTeam = teams[-1]
	copyCell(sheet.cell(8+rowNb, 8), template.cell(tr, 1), rowNb)
	copyCell(sheet.cell(8+rowNb, 9), template.cell(tr, 2), lastTeam[0])
	if colNb == 3:
		thirdStudent = lastTeam[2] if len(lastTeam) == 3 else ''
		copyCell(sheet.cell(8+rowNb, 10), template.cell(tr, 2), lastTeam[1])
		copyCell(sheet.cell(8+rowNb, 11), template.cell(tr, 3), thirdStudent)
	else:
		copyCell(sheet.cell(8+rowNb, 10), template.cell(tr, 3), lastTeam[1])

def write_teams(path,labs):
	if os.path.exists(path):
		os.remove(path)
	workbook = openpyxl.Workbook()

	sheetNb = 0
	for lab in labs:
		if sheetNb == 0:
			worksheet = workbook.worksheets[sheetNb]
			worksheet.title = str(sheetNb+1)
		else:
			workbook.create_sheet(str(sheetNb + 1))
			worksheet = workbook.worksheets[sheetNb]
		write_template(lab,worksheet,sheetNb + 1)
		worksheet.cell(row = 1,column = 1).value = "Équipe"
		worksheet.cell(row = 1,column = 2).value = "Nom 1"
		worksheet.cell(row = 1,column = 3).value = "Nom 2"
		worksheet.cell(row = 1,column = 4).value = "Nom 3"
		rowNb = 2
		for team in lab:
			worksheet.cell(row = rowNb,column = 1).value = rowNb-1
			colNb = 2
			for student in team:
				worksheet.cell(row = rowNb,column = colNb).value = student
				colNb += 1
			rowNb += 1
		sheetNb += 1
	workbook.save(path)

def create_teams(number_of_labs = None, reuse = None):
	folderPath = os.getcwd()
	if number_of_labs is None:
		number_of_labs = easygui.enterbox("How many labs?")
	if number_of_labs is None:
		return
	if reuse is None:
		reuse = easygui.ynbox("Reuse labs already in folder?")
	try:
		if not number_of_labs.isnumeric():
			raise ValueError('Specified number of labs is not a number')
		number_of_labs = int(number_of_labs)
		xlsPath = os.path.join(folderPath, 'etudiants.xlsx')
		if not os.path.exists(xlsPath):
			raise ValueError('Could not find student file at expected location ' + xlsPath)
		students = read_excel(xlsPath)
		pairings = {}
		studentsInTrios = []
		for e in students:
			pairings[e] = []
		labs = []
		reusedLabs = []

		while len(labs) + len(reusedLabs) < number_of_labs:
			teams = []
			if reuse:
				current_lab = len(reusedLabs) + 1
				if os.path.exists(os.path.join(folderPath, 'teams.xlsx')):
					oldTeams,oldStudents = read_teams(os.path.join(folderPath, 'teams.xlsx'),current_lab)
					if oldTeams:
						if set(oldStudents) != set(students):
							raise ValueError('saved labs do not have the same students than current student file')
						teams = oldTeams
				else:
					reuse = False
			if len(teams) != 0:
				for team in teams:
					if len(team) == 3:
						studentsInTrios += team
					for student in team:
						otherStudents = [s for s in team if s is not student]
						pairings[student] += otherStudents
				reusedLabs.append(teams)
				continue
			
			trioTeam = []
			preferredStudentsForTrio = []
			if len(students) - len(studentsInTrios) < 3:
				resetTrios = easygui.ynbox(
					"Too many labs and not enough students, we need them to be in trios a second time. Allow it?")
				if not resetTrios:
					raise ValueError(
						'Too many labs and not enough students to have them participate in only one trio. Aborting.')
				else:
					preferredStudentsForTrio = [s for s in students if s not in studentsInTrios]
					studentsInTrios = []
			remainingStudents = [i for i in students]
			random.shuffle(remainingStudents)
			attempts = 0
			while len(remainingStudents) != 0:
				success = False
				firstStudent = None
				secondStudent = None
				if len(remainingStudents) % 2 != 0: # we build the trio first
					if len(trioTeam) != 0:
						raise ValueError('Programming error : trio team already exists')
					if len(preferredStudentsForTrio) > 2:
						raise ValueError('Programming error : preferredStudentForTrio should have 0, 1 or 2 elements')
					if preferredStudentsForTrio:
						firstStudent = preferredStudentsForTrio.pop()
						remainingStudents.remove(firstStudent)
					else:
						for i in range(len(remainingStudents)):
							if remainingStudents[i] not in studentsInTrios:
								firstStudent = remainingStudents.pop(i)
								break
					if firstStudent is None:
						raise ValueError('All students have already been in trios')
					firstStudentPairings = pairings[firstStudent]
					if preferredStudentsForTrio and preferredStudentsForTrio[0] not in firstStudentPairings:
						secondStudent = preferredStudentsForTrio.pop()
						remainingStudents.remove(secondStudent)
					else:
						for i in range(len(remainingStudents)):
							if remainingStudents[i] not in studentsInTrios and \
									remainingStudents[i] not in firstStudentPairings:
								secondStudent = remainingStudents.pop(i)
								break
					preferredStudentsForTrio = []
					if secondStudent is not None:
						thirdStudent = None
						bothStudentsPairings = firstStudentPairings + pairings[secondStudent]
						for i in range(len(remainingStudents)):
							if remainingStudents[i] not in studentsInTrios and \
									remainingStudents[i] not in bothStudentsPairings:
								thirdStudent = remainingStudents.pop(i)
								break
						if thirdStudent is not None:
							trioTeam = [firstStudent, secondStudent, thirdStudent]
							success = True
				else:
					firstStudent = remainingStudents.pop()
					firstStudentPairings = pairings[firstStudent]
					for i in range(len(remainingStudents)):
						if remainingStudents[i] not in firstStudentPairings:
							teams.append([firstStudent,remainingStudents.pop(i)])
							success = True
							break
				if success:
					attempts = attempts - (attempts % 5) #after a success, we always have 5 attempts before trying something else
				else:
					attempts += 1
					if firstStudent is not None:
						remainingStudents.append(firstStudent)
					if secondStudent is not None:
						remainingStudents.append(secondStudent)
					print('fail ' + str(attempts) + ' with remaining students ' + str(remainingStudents))
					if attempts > 1e5:
						raise ValueError('Found no solution after 100000 iterations. Aborting.')
					elif attempts % 5 == 0:
						#we neeed to pop teams
						random.shuffle(teams)
						milestones = [5,25,125,625,3125,15625,78125]
						for milestone in milestones:
							if attempts % milestone == 0:
								if len(teams) > 0 :
									remainingStudents += teams.pop()
								else:
									if trioTeam:
										remainingStudents += trioTeam
										trioTeam = []
									break
					random.shuffle(remainingStudents)


			#success
			if trioTeam:
				studentsInTrios += trioTeam
				teams.append(trioTeam)
			for team in teams:
				for student in team:
					otherStudents = [s for s in team if s is not student]
					pairings[student] += otherStudents
			labs.append(teams)

		allLabs = reusedLabs + labs
		write_teams(os.path.join(folderPath,'teams.xlsx'),allLabs)
		easygui.msgbox('Teams created successfully, ' + str(len(reusedLabs)) + ' labs were reused.', 'Success')

	except Exception as e:
		easygui.msgbox(str(e),'error')
		raise e


if __name__ == '__main__':
	#create_teams('5',True)
	create_teams()